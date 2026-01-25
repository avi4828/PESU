print('\n')

import openpyxl as op

# Load an existing workbook
wb = op.load_workbook('VideoSales.xlsx')
# print(type(wb))

# call the active worksheet
# ws = wb.active
# print(type(ws),ws)


# call the worksheet by name
# ws1 = wb['SalesData']
# print(type(ws1), ws1)

# count the number of filled rows and columns in a worksheet
ws = wb['SalesData']
# print(ws.max_row)
# print(ws.max_column)



# read the data from specific cell
# print(f'the data stroed in cell D5 is {ws['D5'].value}')

# read the data from singal row but multiple columns
# data = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
# print(data)


# read the data from singal column but multiple rows 
# data = [ws.cell(row=i,column=2).value for i in range(2,ws.max_row+1)]
# print(data)
# for i in data: print(i)


# read the data from a range of cells
data = []
for values in ws.iter_rows(min_row=1,max_row=5,min_col=1,max_col=4,values_only=True):
    data.append(values)
for i in data: print(i)


# display the data in a tabuler manner
# for d1,d2,d3,d4 in data:
#     '''unpack the tupple'''
#     print('{:<10}{:<35}{:<10}{:<5}'.format(d1,d2,d3,d4))

# write to the file
# 1. write in a cell

# ws['K1']='total_sales'
# save the changes in the excel file
# wb.save('VideoSales.xlsx')


# compute a values and write it to the file
# r_pos=2
# c_pos=7
# t_sales=((ws.cell(row=r_pos,column=c_pos).value)+
#         (ws.cell(row=r_pos,column=c_pos+1).value)+
#         (ws.cell(row=r_pos,column=c_pos+2).value)+
#         (ws.cell(row=r_pos,column=c_pos+3).value))
# ws.cell(row=r_pos,column=c_pos+4).value=t_sales
# wb.save('VideoSales.xlsx')


# use a loop ti sum the values in every row
# rpos = 3
# for i in range(3,ws.max_row+1):
#     n_sales = ws.cell(row=rpos,column=7).value
#     s_sales = ws.cell(row=rpos,column=8).value
#     w_sales = ws.cell(row=rpos,column=9).value
#     e_sales = ws.cell(row=rpos,column=10).value
#     total_sales = n_sales + s_sales + w_sales + e_sales
#     ws.cell(row=rpos,column=11).value=total_sales
#     rpos += 1
# wb.save('VideoSales.xlsx')




# nr=(31,'Yoga Master','Trainer	Mobile	2023	Fitness	WellBeing Games	0.85	1.3	0.7	0.9	3.75')
# wb.append('SalesData', nr)

# delete an existing row
# first parameter is position from where the rows have to be deleted 
# decond paramerter is number of rows to be deleted
# ws.delete_rows(ws.max_row,1)
# wb.save('VideoSales.xlsx')


# Avverage()
# ws['M1']='Total Sales'
# ws['M2']='=AVERAGE(K2:K31)'
# wb.save('VideoSales.xlsx')

# counta()
# this count cells that are populated within a specific range
# ws['N1']='Total of populated cells'
# ws['N2']='=COUNTA(E2:E40)'
# wb.save('VideoSales.xlsx')


# countif()
# this counts the number of cells that meet a speific condition
# ws['O1']='number of rows with dports genre'
# ws['O2']='=COUNTIF(E2:E35,"Sports")'
# wb.save('VideoSales.xlsx')

# sumif()
# this sums the cells that satisfy a particular condition
# ws['M5']='Total sports sales'
# ws['M6']='=SUMIF(E2:E33,"Sports",K2:K33)'
# wb.save('VideoSales.xlsx')



# get the name of the active worksheet
# print(ws.title)
# # rename of active worksheet
# ws.title = 'Game Sales Data'
# wb.save('VideoSales.xlsx') 
# print(ws.title)
# ws.title = 'SalesData'
# wb.save('VideoSales.xlsx')
# print(ws.title)


# create a new worksheet
# wb.create_sheet('New Sheet')
# print(wb.sheetnames)
# wb.save('VideoSales.xlsx')


# delete a worksheet
# wb.remove(wb['New Sheet'])
# wb.save('VideoSales.xlsx')


# duplicate a worksheet
# wb.copy_worksheet(wb['SalesData'])
# wb.save('VideoSales.xlsx')
















print('\n')
